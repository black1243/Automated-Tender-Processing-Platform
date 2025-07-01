import openpyxl
from config import OUTPUT_DIR
from przetarg_processor import download_file, extract_text_from_file
from utils import sanitize_filename
import logging
from pathlib import Path
import re
import zipfile
from ai_utils import get_summary_from_ai
from logger_utils import log_action
import py7zr

def extract_all_zips(folder):
    # Always extract all zip/7z files found anywhere in the folder tree into the root folder
    while True:
        found_archives = False
        for path in list(folder.rglob("*")):
            if path.is_file():
                try:
                    if zipfile.is_zipfile(path):
                        print(f"Unzipping ZIP: {path} into {folder}")
                        with zipfile.ZipFile(path, 'r') as zip_ref:
                            zip_ref.extractall(folder)
                        path.unlink()
                        found_archives = True
                    elif path.suffix.lower() == '.7z':
                        print(f"Unzipping 7Z: {path} into {folder}")
                        with py7zr.SevenZipFile(path, mode='r') as archive:
                            archive.extractall(path=folder)
                        path.unlink()
                        found_archives = True
                except Exception as e:
                    print(f"Błąd rozpakowywania {path}: {e}")
        if not found_archives:
            break

def process_tenders(input_excel_file, email_date_prefix=None, output_dir=None):
    if output_dir is None:
        from config import OUTPUT_DIR as DEFAULT_OUTPUT_DIR
        output_dir = DEFAULT_OUTPUT_DIR
    log_action('process_tenders_start', {'input_excel_file': str(input_excel_file)})
    workbook = openpyxl.load_workbook(input_excel_file)
    sheet = workbook.active
    
    # Extract links from column 2 (index 1) that start with https://www.biznes-polska.pl/przetargi/ or https://www.biznes-polska.pl/wyniki-przetargow/
    # Store links with their row index for mapping to tenders
    col2_links_by_row = {}
    if sheet.max_column >= 2:
        col2_cells = sheet['B']  # Column B (index 1)
        for idx, cell in enumerate(col2_cells):
            # Check for plain text links
            if cell.value and isinstance(cell.value, str) and ('https://www.biznes-polska.pl/przetargi/' in cell.value or 'https://www.biznes-polska.pl/wyniki-przetargow/' in cell.value):
                # Extract the full URL
                match = re.search(r'https://www\.biznes-polska\.pl/(przetargi|wyniki-przetargow)/[^\s"]+', cell.value)
                if match:
                    col2_links_by_row[idx+1] = match.group(0)  # Store with 1-based row index
                    log_action('found_col2_link', {'row': idx+1, 'link': match.group(0)})
            # Check for HYPERLINK formulas
            elif cell.value and isinstance(cell.value, str) and 'HYPERLINK' in cell.value.upper():
                # Extract URL from HYPERLINK formula
                match = re.search(r'HYPERLINK\("([^"]+)"', cell.value, re.IGNORECASE)
                if match:
                    url = match.group(1)
                    if 'https://www.biznes-polska.pl/przetargi/' in url or 'https://www.biznes-polska.pl/wyniki-przetargow/' in url:
                        col2_links_by_row[idx+1] = url  # Store with 1-based row index
                        log_action('found_col2_hyperlink_formula', {'row': idx+1, 'link': url})
            # Check for cell hyperlink property
            elif getattr(cell, 'hyperlink', None) and cell.hyperlink.target:
                hlink = cell.hyperlink.target
                if 'https://www.biznes-polska.pl/przetargi/' in hlink or 'https://www.biznes-polska.pl/wyniki-przetargow/' in hlink:
                    col2_links_by_row[idx+1] = hlink  # Store with 1-based row index
                    log_action('found_col2_hyperlink', {'row': idx+1, 'link': hlink})
    
    header_row_index = 10
    headers = [cell.value for cell in sheet[header_row_index]]
    for row_index in range(header_row_index + 1, sheet.max_row + 1):
        row_cells = sheet[row_index]
        # Stop processing if the row is completely empty
        if all((cell.value is None or str(cell.value).strip() == "") for cell in row_cells):
            break
        row_data = {headers[i]: cell.value for i, cell in enumerate(row_cells)}
        if not row_data.get('ID'):
            continue
        organizator = row_data.get('Organizator', 'Nieznany_Organizator')
        termin = row_data.get('Termin składania')
        if hasattr(termin, 'strftime'):
            termin_str = termin.strftime('%Y-%m-%d')
        else:
            termin_str = str(termin).split(' ')[0] if termin else 'Brak_Terminu'
        folder_name = sanitize_filename(f"{organizator}_{termin_str}")
        if email_date_prefix:
            tender_folder = output_dir / email_date_prefix / folder_name
        else:
            tender_folder = output_dir / folder_name
        tender_folder.mkdir(parents=True, exist_ok=True)
        
        # Create link.txt file for this tender if there's a column 2 link for this row
        if row_index in col2_links_by_row:
            link_file_path = tender_folder / 'link.txt'
            with open(link_file_path, 'w', encoding='utf-8') as f:
                f.write(f"{col2_links_by_row[row_index]}\n")
            log_action('tender_link_file_created', {'tender_folder': str(tender_folder), 'link': col2_links_by_row[row_index]})
        
        log_action('tender_processing_start', {'tender_folder': str(tender_folder)})
        # Szukanie linku tylko w kolumnach 28-33
        log_action('searching_attachment_link', {'row_index': row_index, 'tender_folder': str(tender_folder)})
        link = None
        for idx in range(27, 33):  # kolumny 28-33 (indeksy 27-32)
            if idx >= len(row_cells):
                continue
            cell = row_cells[idx]
            if cell.value and isinstance(cell.value, str) and 'https://www.biznes-polska.pl/data/files/' in cell.value:
                match = re.search(r'https://www\\.biznes-polska\\.pl/data/files/[^"\\s]+', cell.value)
                if match:
                    link = match.group(0)
                    break
            if getattr(cell, 'hyperlink', None):
                hlink = cell.hyperlink.target
                if hlink and 'https://www.biznes-polska.pl/data/files/' in hlink:
                    link = hlink
                    break
        if not link:
            log_action('no_attachment_link_found', {'row_index': row_index, 'tender_folder': str(tender_folder)})
        if link:
            log_action('downloading_attachment', {'row_index': row_index, 'tender_folder': str(tender_folder)})
            file_path = download_file(link, tender_folder)
            if file_path and not file_path.suffix.lower() == '.zip':
                new_zip_path = file_path.with_suffix('.zip')
                file_path.rename(new_zip_path)
                file_path = new_zip_path
            if file_path and file_path.suffix.lower() == '.zip':
                try:
                    with zipfile.ZipFile(file_path, 'r') as zip_ref:
                        zip_ref.extractall(tender_folder)
                except Exception as e:
                    pass
            extract_all_zips(tender_folder)
        log_action('generating_summary', {'row_index': row_index, 'tender_folder': str(tender_folder)})
        # Only include .txt files from the _cleaned_txt subfolder
        cleaned_txt_folder = tender_folder / '_cleaned_txt'
        if cleaned_txt_folder.exists() and cleaned_txt_folder.is_dir():
            downloaded_files = [f for f in cleaned_txt_folder.iterdir() if f.is_file() and f.suffix.lower() == '.txt']
        else:
            downloaded_files = []
        files_content = ""
        analyzed_filenames = ""
        for file_path in downloaded_files:
            files_content += f"\n--- {file_path.name} ---\n"
            files_content += extract_text_from_file(file_path)
            analyzed_filenames += f"- {file_path.name}\n"
        summary_content = get_summary_from_ai(row_data, files_content, analyzed_filenames)
        summary_file_path = tender_folder / "_Podsumowanie.md"
        with open(summary_file_path, 'w', encoding='utf-8') as f:
            f.write(summary_content)
        log_action('summary_saved', {'tender_folder': str(tender_folder), 'summary_file': str(summary_file_path)})
        log_action('tender_processing_end', {'tender_folder': str(tender_folder)})
    log_action('process_tenders_end', {'input_excel_file': str(input_excel_file)})
    logging.info("Zakończono przetwarzanie.") 