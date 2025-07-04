import logging
import json
from pathlib import Path
from datetime import datetime, timedelta, date
from gmail_utils import fetch_and_prepare_excel_from_gmail, has_unread_emails
from config import find_excel_file, OUTPUT_DIR
from processing import process_tenders, extract_all_archives
from file_cleaner import convert_and_clean_output_files
import threading
import time
import shutil
from logger_utils import log_action
import zipfile

LOG_FILE = Path('logs.json')

# Persistent log utility
def log_action(action, details):
    log_entry = {
        'timestamp': datetime.now().isoformat(),
        'action': action,
        'details': details
    }
    try:
        if LOG_FILE.exists():
            with open(LOG_FILE, 'r', encoding='utf-8') as f:
                try:
                    logs = json.load(f)
                except Exception:
                    logs = []
        else:
            logs = []
        logs.append(log_entry)
        with open(LOG_FILE, 'w', encoding='utf-8') as f:
            json.dump(logs, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f'Błąd zapisu logów: {e}')

# Helper to parse date input (single, list, or range)
def parse_dates(date_input):
    if not date_input:
        return [None]
    if isinstance(date_input, str):
        return [date_input]
    if isinstance(date_input, list):
        return date_input
    if isinstance(date_input, dict) and 'start' in date_input and 'end' in date_input:
        start = datetime.strptime(date_input['start'], '%Y-%m-%d').date()
        end = datetime.strptime(date_input['end'], '%Y-%m-%d').date()
        return [(start + timedelta(days=i)).isoformat() for i in range((end - start).days + 1)]
    return [None]

# Main orchestrator
def process_emails_and_tenders(date_input=None):
    """
    Processes emails and tenders for given date(s) or range. If None, processes unread emails.
    date_input: None, str (YYYY-MM-DD), list of str, or dict with 'start' and 'end'.
    """
    log_action('bp_1_start_process_emails_and_tenders', {'date_input': date_input})
    dates = parse_dates(date_input)
    processed_files = []
    for date in dates:
        log_action('start_date_processing', {'date': date})
        # Step 1: Fetch and prepare Excel from Gmail, get xlsx path (now always dane/<date>.xlsx)
        xlsx_path = fetch_and_prepare_excel_from_gmail(date, return_xlsx_path=True)
        log_action('bp_2_found_excel', {'date_input': date, 'xlsx_path': str(xlsx_path) if xlsx_path else None})
        if not xlsx_path:
            log_action('bp_2a_no_excel', {'date_input': date})
            continue
        output_dir = OUTPUT_DIR / date
        output_dir.mkdir(parents=True, exist_ok=True)
        log_action('bp_3_created_output_dir', {'output_dir': str(output_dir)})
        # Step 2: Process tenders for this xlsx only in its date folder
        log_action('start_tender_processing', {'date': date, 'xlsx_path': str(xlsx_path), 'output_dir': str(output_dir)})
        try:
            process_tenders_with_logging(xlsx_path, date, output_dir=output_dir)
        except Exception as e:
            log_action('process_tenders_error', {'date': date, 'xlsx_path': str(xlsx_path), 'error': str(e)})
            continue
        log_action('end_tender_processing', {'date': date, 'xlsx_path': str(xlsx_path), 'output_dir': str(output_dir)})
        # Step 3: Clean and convert files to txt in this date folder
        try:
            convert_and_clean_output_files(output_dir=output_dir)
            log_action('clean_files', {'output_dir': str(output_dir), 'date': date})
        except Exception as e:
            log_action('clean_files_error', {'output_dir': str(output_dir), 'date': date, 'error': str(e)})
        # Step 3.5: Generate AI summaries from cleaned txt files
        from ai_utils import get_summary_from_ai
        from utils import sanitize_filename
        for tender_folder in output_dir.iterdir():
            if not tender_folder.is_dir():
                continue
            cleaned_txt_folder = tender_folder / '_cleaned_txt'
            if cleaned_txt_folder.exists() and cleaned_txt_folder.is_dir():
                downloaded_files = [f for f in cleaned_txt_folder.iterdir() if f.is_file() and f.suffix.lower() == '.txt']
            else:
                downloaded_files = []
            files_content = ""
            analyzed_filenames = ""
            for file_path in downloaded_files:
                files_content += f"\n--- {file_path.name} ---\n"
                from przetarg_processor import extract_text_from_file
                files_content += extract_text_from_file(file_path)
                analyzed_filenames += f"- {file_path.name}\n"
            # Try to get row_data from the tender folder name (if possible)
            # If you have a mapping from folder to row_data, use it here. Otherwise, pass an empty dict or minimal info.
            row_data = {}  # TODO: Replace with actual row_data if available
            if downloaded_files:
                summary_content = get_summary_from_ai(row_data, files_content, analyzed_filenames)
                summary_file_path = tender_folder / "_Podsumowanie.md"
                with open(summary_file_path, 'w', encoding='utf-8') as f:
                    f.write(summary_content)
                log_action('summary_saved', {'date': date, 'tender_folder': str(tender_folder), 'summary_file': str(summary_file_path)})
        # Step 4: Log processed files
        for file in output_dir.rglob('*'):
            if file.is_file():
                processed_files.append(str(file))
                log_action('file_processed', {'file': str(file), 'date': date})
        log_action('end_date_processing', {'date': date})
    remove_duplicate_tender_folders()
    log_action('bp_5_end_process_emails_and_tenders', {'date_input': date_input})
    return processed_files

# Utility to get logs
def get_logs():
    if LOG_FILE.exists():
        with open(LOG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

# Startup routine (to be called from main.py or __main__)
def startup_process():
    print('Starting automatic email and tender processing...')
    if has_unread_emails():
        print('Unread emails detected, processing...')
        processed = process_emails_and_tenders()
        print(f'Processed files: {processed}')
    else:
        print('No unread emails detected.')
    print('Startup processing complete.')

# Scheduler for 19:00 and 20:00 automatic checks
def schedule_email_checks():
    def check_loop():
        while True:
            now = datetime.now()
            if now.hour in [19, 20] and now.minute == 0:
                print(f"Automatyczne sprawdzanie maili o {now.hour}:00")
                if has_unread_emails():
                    print('Unread emails detected, processing...')
                    process_emails_and_tenders()
                else:
                    print('No unread emails detected.')
                time.sleep(60)  # Wait a minute to avoid double trigger
            time.sleep(20)  # Check every 20 seconds
    t = threading.Thread(target=check_loop, daemon=True)
    t.start()

def remove_duplicate_tender_folders():
    """
    Skanuje output/*/* i usuwa zdublowane foldery przetargów o tej samej nazwie.
    Zostawia tylko pierwszy napotkany folder o danej nazwie.
    """
    seen = {}
    for date_dir in OUTPUT_DIR.iterdir():
        if not date_dir.is_dir():
            continue
        for tender_dir in date_dir.iterdir():
            if not tender_dir.is_dir():
                continue
            if tender_dir.name not in seen:
                seen[tender_dir.name] = tender_dir
            else:
                print(f"Usuwam duplikat: {tender_dir}")
                shutil.rmtree(tender_dir)

def move_tenders_to_date_folder(temp_output_subdir, date):
    date_folder = OUTPUT_DIR / date
    date_folder.mkdir(parents=True, exist_ok=True)
    for item in temp_output_subdir.iterdir():
        if item.is_dir():
            # Nie przenoś folderu, który ma nazwę identyczną jak data
            if item.name == date:
                # Przenieś zawartość tego folderu do date_folder
                for subitem in item.iterdir():
                    target = date_folder / subitem.name
                    if target.exists():
                        shutil.rmtree(target)
                    shutil.move(str(subitem), str(target))
                # Usuń pusty podfolder
                try:
                    item.rmdir()
                except Exception:
                    pass
                continue
            target = date_folder / item.name
            if target.exists():
                shutil.rmtree(target)
            shutil.move(str(item), str(target))
    # Usuń pusty tymczasowy podfolder
    try:
        temp_output_subdir.rmdir()
    except Exception:
        pass

# Wrap process_tenders to add logging for each row and attachment
from processing import process_tenders as _process_tenders

def process_tenders_with_logging(input_excel_file, email_date_prefix=None, output_dir=None):
    import openpyxl
    from przetarg_processor import download_file, extract_text_from_file
    from utils import sanitize_filename
    import logging
    import re
    import zipfile
    from pathlib import Path
    workbook = openpyxl.load_workbook(input_excel_file)
    sheet = workbook.active
    
    # Extract links from column 2 (index 1) that start with https://www.biznes-polska.pl/przetargi/ or https://www.biznes-polska.pl/wyniki-przetargow/
    # Store links with their row index for mapping to tenders
    col2_links_by_row = {}
    if sheet.max_column >= 2:
        col2_cells = sheet['B']  # Column B (index 1)
        for idx, cell in enumerate(col2_cells):
            # Check for plain text links
            if cell.value and isinstance(cell.value, str) and ('https://www.biznes-polska.pl/przetargi/' in cell.value or 'https://www.biznes-polska.pl/wyniki-przetargow/' in cell.value):
                # Extract the full URL
                match = re.search(r'https://www\.biznes-polska\.pl/(przetargi|wyniki-przetargow)/[^\s"]+', cell.value)
                if match:
                    col2_links_by_row[idx+1] = match.group(0)  # Store with 1-based row index
                    log_action('found_col2_link', {'date': email_date_prefix, 'row': idx+1, 'link': match.group(0)})
            # Check for HYPERLINK formulas
            elif cell.value and isinstance(cell.value, str) and 'HYPERLINK' in cell.value.upper():
                # Extract URL from HYPERLINK formula
                match = re.search(r'HYPERLINK\("([^"]+)"', cell.value, re.IGNORECASE)
                if match:
                    url = match.group(1)
                    if 'https://www.biznes-polska.pl/przetargi/' in url or 'https://www.biznes-polska.pl/wyniki-przetargow/' in url:
                        col2_links_by_row[idx+1] = url  # Store with 1-based row index
                        log_action('found_col2_hyperlink_formula', {'date': email_date_prefix, 'row': idx+1, 'link': url})
            # Check for cell hyperlink property
            elif getattr(cell, 'hyperlink', None) and cell.hyperlink.target:
                hlink = cell.hyperlink.target
                if 'https://www.biznes-polska.pl/przetargi/' in hlink or 'https://www.biznes-polska.pl/wyniki-przetargow/' in hlink:
                    col2_links_by_row[idx+1] = hlink  # Store with 1-based row index
                    log_action('found_col2_hyperlink', {'date': email_date_prefix, 'row': idx+1, 'link': hlink})
    
    header_row_index = 10
    headers = [cell.value for cell in sheet[header_row_index]]
    for row_index in range(header_row_index + 1, sheet.max_row + 1):
        row_cells = sheet[row_index]
        # Stop processing if the first column 'ID' is not a number
        id_value = row_cells[0].value if len(row_cells) > 0 else None
        try:
            if id_value is None or str(id_value).strip() == '' or not str(id_value).strip().isdigit():
                break
        except Exception:
            break
        row_data = {headers[i]: cell.value for i, cell in enumerate(row_cells)}
        log_action('row_data', {'date': email_date_prefix, 'row_index': row_index, 'row_data': str(row_data)})
        if not row_data.get('ID'):
            log_action('skip_row_no_id', {'date': email_date_prefix, 'row_index': row_index})
            continue
        organizator = row_data.get('Organizator', 'Nieznany_Organizator')
        termin = row_data.get('Termin składania')
        # --- SKIP IF DEADLINE HAS PASSED ---
        deadline_date = None
        if hasattr(termin, 'date'):
            deadline_date = termin.date() if hasattr(termin, 'date') else termin
        elif hasattr(termin, 'strftime'):
            deadline_date = termin.date() if hasattr(termin, 'date') else termin
        elif isinstance(termin, str):
            try:
                deadline_date = datetime.strptime(termin.split(' ')[0], '%Y-%m-%d').date()
            except Exception:
                deadline_date = None
        if deadline_date and deadline_date < date.today():
            log_action('skip_row_deadline_passed', {'date': email_date_prefix, 'row_index': row_index, 'deadline': str(deadline_date)})
            continue
        # --- END SKIP ---
        if hasattr(termin, 'strftime'):
            termin_str = termin.strftime('%Y-%m-%d')
        else:
            termin_str = str(termin).split(' ')[0] if termin else 'Brak_Terminu'
        folder_name = sanitize_filename(f"{organizator}_{termin_str}")
        if email_date_prefix:
            tender_folder = output_dir / folder_name
        else:
            tender_folder = output_dir / folder_name
        tender_folder.mkdir(parents=True, exist_ok=True)
        
        # Create link.txt file for this tender if there's a column 2 link for this row
        if row_index in col2_links_by_row:
            link_file_path = tender_folder / 'link.txt'
            with open(link_file_path, 'w', encoding='utf-8') as f:
                f.write(f"{col2_links_by_row[row_index]}\n")
            log_action('tender_link_file_created', {'date': email_date_prefix, 'tender_folder': str(tender_folder), 'link': col2_links_by_row[row_index]})
        
        downloaded_files = []
        link = None
        for idx, cell in enumerate(row_cells):
            log_action('cell_value', {'date': email_date_prefix, 'row_index': row_index, 'col': idx+1, 'value': cell.value, 'hyperlink': getattr(cell, 'hyperlink', None)})
        for idx, cell in enumerate(row_cells):
            if cell.value and isinstance(cell.value, str) and 'https://www.biznes-polska.pl/data/files/' in cell.value:
                match = re.search(r'https://www\.biznes-polska\.pl/data/files/[^"\s]+', cell.value)
                if match:
                    link = match.group(0)
                    log_action('found_link', {'date': email_date_prefix, 'row_index': row_index, 'col': idx+1, 'link': link})
                    break
            if getattr(cell, 'hyperlink', None):
                hlink = cell.hyperlink.target
                if hlink and 'https://www.biznes-polska.pl/data/files/' in hlink:
                    link = hlink
                    log_action('found_link_hyperlink', {'date': email_date_prefix, 'row_index': row_index, 'col': idx+1, 'link': link})
                    break
        if not link:
            log_action('no_link_found', {'date': email_date_prefix, 'row_index': row_index})
        if link:
            file_path = download_file(link, tender_folder)
            log_action('download_file', {'date': email_date_prefix, 'row_index': row_index, 'link': link, 'file_path': str(file_path)})
            if file_path and not file_path.suffix.lower() == '.zip':
                new_zip_path = file_path.with_suffix('.zip')
                file_path.rename(new_zip_path)
                file_path = new_zip_path
                log_action('rename_to_zip', {'date': email_date_prefix, 'row_index': row_index, 'file_path': str(file_path)})
            if file_path and file_path.suffix.lower() == '.zip':
                try:
                    with zipfile.ZipFile(file_path, 'r') as zip_ref:
                        zip_ref.extractall(tender_folder)
                    log_action('extract_zip', {'date': email_date_prefix, 'row_index': row_index, 'file_path': str(file_path), 'tender_folder': str(tender_folder)})
                    file_path.unlink()
                except Exception as e:
                    log_action('extract_zip_error', {'date': email_date_prefix, 'row_index': row_index, 'file_path': str(file_path), 'error': str(e)})
            # Extract all nested archives (ZIP, 7Z, RAR) recursively
            extract_all_archives(tender_folder)
            # Optionally, log all files in tender_folder after extraction
            for extracted in tender_folder.iterdir():
                if extracted.is_file() and extracted.suffix.lower() in ['.pdf', '.docx', '.doc', '.txt']:
                    downloaded_files.append(extracted)
                    log_action('file_for_analysis', {'date': email_date_prefix, 'row_index': row_index, 'file': str(extracted)})
        
        # Only include .txt files from the _cleaned_txt subfolder
        cleaned_txt_folder = tender_folder / '_cleaned_txt'
        if cleaned_txt_folder.exists() and cleaned_txt_folder.is_dir():
            downloaded_files = [f for f in cleaned_txt_folder.iterdir() if f.is_file() and f.suffix.lower() == '.txt']
        else:
            downloaded_files = []
        files_content = ""
        analyzed_filenames = ""
        for file_path in downloaded_files:
            files_content += f"\n--- {file_path.name} ---\n"
            files_content += extract_text_from_file(file_path)
            analyzed_filenames += f"- {file_path.name}\n"
        if downloaded_files:  # Only generate summary if there are files
            summary_content = get_summary_from_ai(row_data, files_content, analyzed_filenames)
            summary_file_path = tender_folder / "_Podsumowanie.md"
            with open(summary_file_path, 'w', encoding='utf-8') as f:
                f.write(summary_content)
            log_action('summary_saved', {'date': email_date_prefix, 'row_index': row_index, 'summary_file': str(summary_file_path)})
    
    log_action('end_process_tenders_with_logging', {'date': email_date_prefix, 'xlsx': str(input_excel_file)}) 