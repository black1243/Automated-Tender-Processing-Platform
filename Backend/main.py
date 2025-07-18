import openpyxl
import xlrd
from pathlib import Path
import os
from dotenv import load_dotenv
from config import OUTPUT_DIR, find_excel_file
from przetarg_processor import download_file, extract_text_from_file
from utils import sanitize_filename
import zipfile
import py7zr
import logging
import re
import base64
import requests
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
import io
from gmail_utils import fetch_and_prepare_excel_from_gmail
from processing import process_tenders
from ai_utils import get_summary_from_ai

load_dotenv()

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s: %(message)s')

def extract_hyperlink_from_formula(cell_value):
    # Obsługuje polskie i angielskie wersje formuły
    match = re.match(r'=HYPERŁĄCZE\("([^"]+)";"[^"]+"\)', cell_value, re.IGNORECASE)
    if not match:
        match = re.match(r'=HYPERLINK\("([^"]+)";"[^"]+"\)', cell_value, re.IGNORECASE)
    if match:
        return match.group(1)
    return None

def convert_xls_to_xlsx(xls_path):
    """
    Converts .xls to .xlsx, preserving hyperlinks (including those from HYPERLINK formulas).
    """
    xlsx_path = Path(str(xls_path).replace('.xls', '.xlsx'))
    try:
        book = xlrd.open_workbook(xls_path, formatting_info=False)
        sheet = book.sheet_by_index(0)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet.name
        hyperlink_re = re.compile(r'=HYPER(?:LINK|ŁĄCZE)\("([^"]+)";?"?([^"]*)"?\)', re.IGNORECASE)
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                value = cell.value
                cell_ref = f"{get_column_letter(col_idx+1)}{row_idx+1}"
                if isinstance(value, str) and ("HYPERLINK" in value.upper() or "HIPERŁĄCZE" in value.upper()):
                    match = hyperlink_re.match(value)
                    if match:
                        url = match.group(1)
                        display = match.group(2) if match.group(2) else url
                        ws[cell_ref].value = display
                        ws[cell_ref].hyperlink = url
                        ws[cell_ref].font = Font(color="0000FF", underline="single")
                    else:
                        ws[cell_ref].value = value
                else:
                    ws[cell_ref].value = value
        wb.save(xlsx_path)
        return xlsx_path
    except Exception as e:
        print(f"Błąd konwersji pliku .xls na .xlsx (z hyperlinkami): {e}")
        return None

def process_xls_file(xls_path, output_dir):
    print(f"Przetwarzam plik .xls: {xls_path}")
    book = xlrd.open_workbook(xls_path, formatting_info=True)
    sheet = book.sheet_by_index(0)
    
    # Extract links from column 2 (index 1) that start with https://www.biznes-polska.pl/przetargi/ or https://www.biznes-polska.pl/wyniki-przetargow/
    # Store links with their row index for mapping to tenders
    col2_links_by_row = {}
    if sheet.ncols >= 2:
        for row_idx in range(sheet.nrows):
            cell_value = sheet.cell_value(row_idx, 1)  # Column 2 (index 1)
            # Check for plain text links
            if isinstance(cell_value, str) and ('https://www.biznes-polska.pl/przetargi/' in cell_value or 'https://www.biznes-polska.pl/wyniki-przetargow/' in cell_value):
                # Extract the full URL
                match = re.search(r'https://www\.biznes-polska\.pl/(przetargi|wyniki-przetargow)/[^\s"]+', cell_value)
                if match:
                    col2_links_by_row[row_idx+1] = match.group(0)  # Store with 1-based row index
                    logging.info(f"Znaleziono link w kolumnie 2, wiersz {row_idx+1}: {match.group(0)}")
            # Check for HYPERLINK formulas
            elif isinstance(cell_value, str) and 'HYPERLINK' in cell_value.upper():
                # Extract URL from HYPERLINK formula
                match = re.search(r'HYPERLINK\("([^"]+)"', cell_value, re.IGNORECASE)
                if match:
                    url = match.group(1)
                    if 'https://www.biznes-polska.pl/przetargi/' in url or 'https://www.biznes-polska.pl/wyniki-przetargow/' in url:
                        col2_links_by_row[row_idx+1] = url  # Store with 1-based row index
                        logging.info(f"Znaleziono link w formule HYPERLINK w kolumnie 2, wiersz {row_idx+1}: {url}")
    
    header_row_index = 9  # xlrd: 0-based, openpyxl: 1-based
    headers = [sheet.cell_value(header_row_index, col) for col in range(sheet.ncols)]
    for row_idx in range(header_row_index + 1, sheet.nrows):
        row_values = [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]
        row_data = dict(zip(headers, row_values))
        if not row_data.get('ID'):
            continue
        organizator = row_data.get('Organizator', 'Nieznany_Organizator')
        termin = row_data.get('Termin składania')
        if hasattr(termin, 'strftime'):
            termin_str = termin.strftime('%Y-%m-%d')
        else:
            termin_str = str(termin).split(' ')[0] if termin else 'Brak_Terminu'
        folder_name = sanitize_filename(f"{organizator}_{termin_str}")
        tender_folder = output_dir / folder_name
        tender_folder.mkdir(parents=True, exist_ok=True)
        
        # Create link.txt file for this tender if there's a column 2 link for this row
        if row_idx+1 in col2_links_by_row:  # row_idx is 0-based, so add 1 to match 1-based indexing
            link_file_path = tender_folder / 'link.txt'
            with open(link_file_path, 'w', encoding='utf-8') as f:
                f.write(f"{col2_links_by_row[row_idx+1]}\n")
            logging.info(f"Utworzono plik link.txt dla {folder_name}: {col2_links_by_row[row_idx+1]}")
        
        downloaded_files = []
        link = None
        # Szukaj linku w każdej komórce wiersza (formuła HYPERLINK/HIPERŁĄCZE)
        for col_idx, cell in enumerate(sheet.row(row_idx)):
            val = cell.value
            if isinstance(val, str) and ('HYPERLINK' in val.upper() or 'HIPERŁĄCZE' in val.upper()):
                match = re.search(r'HYPER(?:LINK|ŁĄCZE)\("([^"]+)"', val, re.IGNORECASE)
                if match:
                    link = match.group(1)
                    logging.info(f"Znaleziono link w formule w kolumnie {col_idx+1}: {link}")
                    break
            if isinstance(val, str) and 'https://www.biznes-polska.pl/data/files/' in val:
                match = re.search(r'https://www\.biznes-polska\.pl/data/files/[^\s\"]+', val)
                if match:
                    link = match.group(0)
                    logging.info(f"Znaleziono link w tekście w kolumnie {col_idx+1}: {link}")
                    break
        if not link:
            logging.info(f"Brak linku do załącznika w całym wierszu {row_idx+1}.")
        if link:
            file_path = download_file(link, tender_folder)
            if file_path and not file_path.suffix.lower() == '.zip':
                new_zip_path = file_path.with_suffix('.zip')
                file_path.rename(new_zip_path)
                file_path = new_zip_path
                logging.info(f"Zmieniono rozszerzenie pobranego pliku na .zip: {file_path}")
            if file_path and file_path.suffix.lower() == '.zip':
                try:
                    with zipfile.ZipFile(file_path, 'r') as zip_ref:
                        zip_ref.extractall(tender_folder)
                    logging.info(f"Rozpakowano ZIP: {file_path} do {tender_folder}")
                    for extracted in tender_folder.iterdir():
                        if extracted.is_file():
                            if extracted.suffix.lower() == '.7z':
                                # Extract .7z files
                                try:
                                    with py7zr.SevenZipFile(extracted, mode='r') as archive:
                                        archive.extractall(path=tender_folder)
                                    logging.info(f"Rozpakowano 7Z: {extracted} do {tender_folder}")
                                    # Remove the .7z file after extraction
                                    extracted.unlink()
                                except Exception as e:
                                    logging.error(f"Błąd podczas rozpakowywania 7Z {extracted}: {e}")
                            elif extracted.suffix.lower() in ['.pdf', '.docx', '.doc', '.txt']:
                                downloaded_files.append(extracted)
                                logging.info(f"Dodano do analizy: {extracted}")
                    
                    # After all extractions, scan again for newly extracted files
                    for extracted in tender_folder.iterdir():
                        if extracted.is_file() and extracted.suffix.lower() in ['.pdf', '.docx', '.doc', '.txt']:
                            if extracted not in downloaded_files:
                                downloaded_files.append(extracted)
                                logging.info(f"Dodano do analizy (po rozpakowaniu 7Z): {extracted}")
                except Exception as e:
                    logging.error(f"Błąd podczas rozpakowywania ZIP: {e}")
            elif file_path:
                downloaded_files.append(file_path)
                logging.info(f"Dodano do analizy: {file_path}")
        files_content = ""
        analyzed_filenames = ""
        for file_path in downloaded_files:
            files_content += f"\n--- {file_path.name} ---\n"
            files_content += extract_text_from_file(file_path)[:2000]
            analyzed_filenames += f"- {file_path.name}\n"
        summary_content = get_summary_from_ai(row_data, files_content, analyzed_filenames)
        summary_file_path = tender_folder / "_Podsumowanie.md"
        with open(summary_file_path, 'w', encoding='utf-8') as f:
            f.write(summary_content)
        print(f"Zapisano podsumowanie w: {summary_file_path}")

def convert_xls_to_xlsx_via_gdrive(xls_path, output_dir):
    """
    Uploads .xls to Google Drive, converts to Google Sheets, exports as .xlsx, and downloads to output_dir.
    Returns the path to the downloaded .xlsx file or None on failure.
    """
    creds = get_google_credentials(SCOPES)
    drive_service = build('drive', 'v3', credentials=creds)
    # Upload the .xls file
    file_metadata = {'name': xls_path.name, 'mimeType': 'application/vnd.google-apps.spreadsheet'}
    media = MediaFileUpload(str(xls_path), mimetype='application/vnd.ms-excel')
    uploaded = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    file_id = uploaded.get('id')
    if not file_id:
        print("Nie udało się przesłać pliku na Google Drive.")
        return None
    # Export as .xlsx
    request = drive_service.files().export_media(fileId=file_id, mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    xlsx_path = output_dir / (xls_path.stem + '.xlsx')
    fh = io.FileIO(xlsx_path, 'wb')
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
    fh.close()
    # Clean up: delete the file from Drive
    drive_service.files().delete(fileId=file_id).execute()
    print(f"Pobrano przekonwertowany plik: {xlsx_path}")
    return xlsx_path

def main():
    # Step 1: Fetch file from Gmail (if needed)
    # fetch_and_prepare_excel_from_gmail(selected_date)  # Call from UI or CLI

    # Step 2: Read email date for output structure
    email_date_file = Path('dane/email_date.txt')
    email_date_prefix = email_date_file.read_text(encoding='utf-8').strip() if email_date_file.exists() else None

    # Step 3: Find and process the Excel file
    input_excel_file = find_excel_file()
    if not input_excel_file:
        logging.error("Brak pliku Excel w folderze dane!")
        return

    process_tenders(input_excel_file, email_date_prefix)

    # Step 4: Clean up
    if email_date_file.exists():
        email_date_file.unlink()

if __name__ == '__main__':
    from backend_processor import startup_process, schedule_email_checks
    schedule_email_checks()
    startup_process() 